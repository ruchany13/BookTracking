from openpyxl import Workbook,load_workbook
#Dosya oluşturma
def dosya_acma():
    wb=Workbook()
    ws=wb.active
    ws.append([1,2,3])
    wb.save("deneme.xlsx")
dosya_acma()  
#Dosyayı seçme:
#ws=wb.active
#print(ws)

##------Çalışma sayfası seçimi-------
wb = load_workbook("isimler.xlsx")
ws = wb.sheetnames
#print(ws)   # ['İsimler', 'Şehirler']
#ws = wb["Şehirler"]
#print(ws)   # <Worksheet "Şehirler">
#-------------------------------------

#dosyaya yazı ekleme
def ekeleme():
    wb=Workbook()
    ws=wb.active
    ws['A1']="İlk satır"
    ws['B3']="B3"
    ws.append(["1",2,3,4]) # sıradaki satıra ekleme yapıyor 
    wb.save("deneme.xlsx") #Kaydetmeyi atlama

#Dosyayı okuma iki yöntem var:

#print(ws['A4'].value)  #1.Yöntem

#print(ws.cell(4,1).value)  #2. Yöntem

#Çoklu dosya okuma yöntemleri:

#ekeleme()

def okuma():
    wb=Workbook()
    ws=wb.active
    for satir in range (1,5):
        for sutun in range (1,5):
            print(" | "+str(ws.cell(satir,sutun).value)+" |",end="")
        print()
    
#Hepsini yazdır:
    for satir in range(1,ws.max_row+1):
        for sutun in range(1,ws.max_column+1):
            a=" | " + str(ws.cell(satir,sutun).value) + " : "
    


#Değiştirme için aynı satıra ekleme yapmak lazım.





