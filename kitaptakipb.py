from openpyxl import Workbook, load_workbook

# Kitapların bilgilerini durumuna göre excel'deki çalışma alanlarına ekliyoruz.(Okunmuş,okunuyor,okunacak)
# Kitabın numarası ile sutunu buluyoruz ve uygun değişikliği yapıyoruz.
# kitab numarası:2   değşiklik: sayfa ekleme  işlem satırı: E2 bazen de (2,4) gibi
# 1) Okunuyor  2)Okunacak  3)Okundu  4)Biten

wb = load_workbook("kitaptakip.xlsx")
ws = wb.active


# Kitapları ekleme için kullanıcıdan veri alma fonksiyonu
def kitap_bilgileri(durum):
    kitap = input("Kitabın adı: ")
    yazar = input("Kitabın yazarı: ")
    sayfa = int(input("Kitabın sayfa sayısı: "))
    if durum == "okunuyor":
        baslangic = input("Başlangıç tarihini giriniz: ")
        okunan = int(input("Okuduysanız sayfa sayısı yoksa '0' yazınız: "))
        return kitap, yazar, sayfa, okunan, baslangic
    elif durum == "okundu":
        baslangic = input("Başlangıç tarihini giriniz.")
        bitis = input("Bitiş tarihini yazınız.")
        return bitis, kitap, yazar, sayfa, baslangic, bitis
    elif durum == "okunacak":
        return kitap, yazar, sayfa


# Kullanıcın sistemde takip etiiği kitabın otomatik olarak okunan a ekleme fonksiyonu
def yeni_kitap_biten(kosul, numara):
    # Bitmiş olan kitabı eklemek için bu taraf.
    if kosul == "bitti":
        ws = wb['Okunan']
        kitap = ws.cell(numara, 1).value
        yazar = ws.cell(numara, 2).value
        sayfa = ws.cell(numara, 3).value

        # Biten kitabın bilgilerini akatrıyoruz
        ws = wb['Biten']
        ws.append([kitap, yazar, sayfa])
        wb.save("kitaptakip.xlsx")


    # Yeni eklenecek olan okuduklarım,okumak istediklerim için burası.
    elif kosul == "yeni":  # TODO yeni_kitap_okunan ile burası için birleşme veya veriler için fonksiyon olabilir.

        print(ws)


# Okunmakta olan kitabın sayfasını güncelleme
def güncel_kitap():
    ws = wb['Okunan']
    print(ws)
    a = 1
    for satir in range(1, ws.max_row + 1):
        for sutun in range(1, ws.max_column + 1):
            print("" + str(ws.cell(satir, sutun).value) + "\t", end=":")

            print()
        print("\t", a, "\n")
        a += 1

    numara = int(input("Kitap numarasını yazınız:"))
    numara += 1
    # n: okunan sayfa satırı / n2 : kitabın toplam sayfasının satırı 
    okunan = int(input("En son okuduğunuz sayfa:"))
    n, n2 = "D" + str(numara), "C" + str(numara)

    önce_okunan = ws[n].value
    ws[n] = okunan
    sayfa = ws[n2].value

    print("Bugün {} sayfa kitap okudunuz.".format(okunan - önce_okunan))

    # Kitap bittiğinde bitenlere ekleme yapma bölümü

    if sayfa == okunan:
        print("Tebrikler kitabınızı hemen 'bitenler' kategorisine alıyoruz..")
        yeni_kitap_biten("bitti", numara)
    else:
        wb.save("kitaptakip.xlsx")


# Kullanıcının sisteme ekliyeceği bölüm
def yeni_kitap_okunan():  # yeni_kitap_biten ile burası için birleşme veya veriler için fonksiyon olabilir.
    ws = wb["Okunan"]
    print(ws)

    kitap, yazar, sayfa, okunan, baslangic = kitap_bilgileri("okunuyor")

    ws.append([kitap, yazar, sayfa, okunan, baslangic])
    wb.save("kitaptakip.xlsx")


# Kullanıcın okumak istediği kitapları "Okunacak"'a ekleyen fonksiyon 4
def yeni_kitap_okunacak():
    ws = wb["Okunacak"]
    print(ws)
    kitap, yazar, sayfa = kitap_bilgileri("okunacak")
    ws.append([kitap, yazar, sayfa])


while True:
    giris = input("""
------------------------------------------------------------------------
           |     Kitap takip uygulamasına Hoşgeldiniz. Version 1.0      |
           |    Ruchan Yalçın                                           |
------------------------------------------------------------------------                
                    Şu anda okuduğunuz veya başlıyacağınız kitap   : 1
                    
                    Okumakta olduğunuz kitabın sayfasını güncelleme: 2
                    
                    Kitap listesine önceden okunanı eklemek için   : 3
                    
                    Okunacaklar listesine kitap eklemek için       : 4
                    
                    tuşuna basınız: """)
    if giris == "1":
        yeni_kitap_okunan()
    elif giris == "2":
        güncel_kitap()
    elif giris == "3":
        yeni_kitap_biten("bitti", 0)
    elif giris == "4":
        yeni_kitap_okunacak()
    elif giris == "q":
        break
