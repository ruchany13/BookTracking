from openpyxl import Workbook
wb = Workbook() #Örnekliyoruz tahminimce
ws = wb.active 
ws = wb.create_sheet("Deneme") #Yeni bir çalışma sayfası oluşturuyor.
ws3 = wb.create_sheet("Çalışma")
ws.title="Bu ne" #Bu seçili çalışma sayfasının ismini değiştiryor

ws['A1'] = "İlk Satır"
# ws 1,2,3 vb. şekilde çalışma sayfasını oluştur
#Okuma veya ekleme durumunda çalışma sayfasını direkt yaz(ws1,ws2 vb.)


ws2=wb["Bu ne"]
ws3['A2']="sON sASADLSL" # Oluşturduğumuz her çalışma sayfasının adı ile erişebiliyoruz.
ws['A2'] = "Merhaba"
ws['C9']="hello world"
for row in range(1, 40):
    ws3.append(range(600))# Ekleme işi

for row in range(1,ws.max_row+1):
    for column in range(1,ws.max_column+1)
     

wb.save("a.xlsx")

